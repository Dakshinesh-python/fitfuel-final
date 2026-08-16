"""
generate_reports.py

Runs AFTER `pytest` finishes (see .github/workflows/android-e2e.yml).
Reads the single reports/execution-results.json that conftest.py's
pytest_sessionfinish() already wrote, and produces:

  reports/Automation_Test_Report.xlsx - Executed Tests / Passed / Failed /
                                          Skipped / Execution Metrics / Defect
                                          Summary sheets
  reports/execution-report.html       - full per-test table
  reports/dashboard.html              - pass-rate gate + per-module bar chart
  reports/summary.md                  - short markdown summary for a PR comment

Deliberately simpler than selenium-tests/generate_reports.py in one
respect: that suite merges multiple result_*.json shards because it runs
under pytest-xdist across several browser workers. This suite runs a
single Appium/Flutter-driver session (one physical/virtual device, no
meaningful way to parallelize an app install + Observatory handshake
across workers) so there is only ever one result file, written directly
by conftest.py rather than assembled here. Everything downstream --
sheet layout, HTML structure, gate logic, exit code -- is kept identical
to selenium-tests/ on purpose, so a reviewer opening either report gets
the same format.
"""
from __future__ import annotations

import json
import os
import sys
import time

sys.path.insert(0, os.path.dirname(__file__))

import config  # noqa: E402

EXPECTED_TEST_COUNT = int(os.environ.get("EXPECTED_TEST_COUNT", "402"))

MODULE_SEVERITY = {
    "test_00_ui_chrome": "LOW",
    "test_01_authentication": "HIGH",
    "test_02_registration": "HIGH",
    "test_03_health_assessment": "HIGH",
    "test_04_navigation": "MEDIUM",
    "test_05_dashboard": "MEDIUM",
    "test_06_recommendations": "MEDIUM",
    "test_07_meal_plan": "MEDIUM",
    "test_08_progress_crud": "HIGH",
    "test_09_chat": "MEDIUM",
    "test_10_profile_settings": "MEDIUM",
    "test_11_input_validation": "MEDIUM",
    "test_12_error_handling": "MEDIUM",
    "test_13_session_management": "HIGH",
    "test_14_accessibility": "LOW",
    "test_15_responsiveness": "LOW",
    "test_16_form_field_matrices": "MEDIUM",
    "test_17_extended_matrices": "LOW",
    "test_18_integration_regression": "HIGH",
    "test_19_additional_coverage": "MEDIUM",
}

STATUS_COLORS = {
    "PASSED": "#1e8e4a",
    "FAILED": "#d64545",
    "SKIPPED": "#b58900",
    "XFAIL": "#6b7280",
}


def load_results() -> dict:
    path = os.path.join(config.REPORTS_DIR, "execution-results.json")
    if not os.path.exists(path):
        print(f"{path} not found -- did pytest run first?", file=sys.stderr)
        sys.exit(1)
    with open(path, encoding="utf-8") as f:
        payload = json.load(f)

    total = payload["summary"]["total"]
    print(f"Loaded {total} test results from {path}")
    if payload.get("partial"):
        print(
            f"WARNING: this results file is PARTIAL -- pytest_sessionfinish never ran, "
            f"meaning the run was interrupted (job/step timeout, OOM kill, crashed "
            f"Appium/emulator session, etc.) after {total} test(s) recorded, not because "
            f"only {total} test(s) exist. Check reports/logs/ for how far the run actually "
            f"got and the pytest job log for what killed it -- do NOT read this as "
            f"'{total} tests ran and that's the real result'.",
            file=sys.stderr,
        )
    elif total != EXPECTED_TEST_COUNT:
        print(
            f"WARNING: expected {EXPECTED_TEST_COUNT} test results, got {total}. "
            f"This usually means -k/-m filters were applied, or some tests were skipped "
            f"at collection. Check the pytest job log.",
            file=sys.stderr,
        )
    return payload


def short_test_id(nodeid: str) -> str:
    return nodeid.split("::", 1)[-1]


def write_excel_report(payload: dict):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    results = payload["results"]
    s = payload["summary"]
    generated_at = payload["generated_at"]

    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    header_font = Font(bold=True)
    status_fill = {
        "PASSED": PatternFill(start_color="EAF7EE", end_color="EAF7EE", fill_type="solid"),
        "FAILED": PatternFill(start_color="FDECEB", end_color="FDECEB", fill_type="solid"),
        "SKIPPED": PatternFill(start_color="FDF3D9", end_color="FDF3D9", fill_type="solid"),
        "XFAIL": PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"),
    }

    def autosize(ws):
        for col_cells in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, 10), 90)

    def header_row(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

    ws = wb.active
    ws.title = "Executed Tests"
    header_row(ws, ["#", "Test ID", "Module", "Markers", "Status", "Duration (s)"])
    for i, r in enumerate(results, start=1):
        ws.append(
            [i, short_test_id(r["nodeid"]), r["module_name"], ", ".join(r.get("markers", [])), r["status"], r["duration_s"]]
        )
        ws.cell(row=i + 1, column=5).fill = status_fill.get(r["status"], PatternFill())
    autosize(ws)

    for status_key, sheet_name in (("PASSED", "Passed"), ("FAILED", "Failed"), ("SKIPPED", "Skipped")):
        ws2 = wb.create_sheet(sheet_name)
        header_row(ws2, ["#", "Test ID", "Module", "Duration (s)"])
        rows = [r for r in results if r["status"] == status_key]
        for i, r in enumerate(rows, start=1):
            ws2.append([i, short_test_id(r["nodeid"]), r["module_name"], r["duration_s"]])
        autosize(ws2)

    ws3 = wb.create_sheet("Execution Metrics")
    header_row(ws3, ["Metric", "Value"])
    ws3.append(["Run At", generated_at])
    ws3.append(["App Package", payload["app_under_test"]["package"]])
    ws3.append(["Backend Base URL", payload["app_under_test"]["backend_base_url"]])
    ws3.append(["Platform", payload["app_under_test"]["platform"]])
    ws3.append(["Platform Version", payload["app_under_test"]["platform_version"]])
    ws3.append(["Total Tests", s["total"]])
    ws3.append(["Expected Total", EXPECTED_TEST_COUNT])
    ws3.append(["Passed", s["passed"]])
    ws3.append(["Failed", s["failed"]])
    ws3.append(["Skipped", s["skipped"]])
    ws3.append(["Pass Rate (%)", s["pass_rate"]])
    ws3.append(["Gate Threshold (%)", s["gate_threshold_pct"]])
    ws3.append(["Gate Passed", "YES" if s["gate_passed"] else "NO"])
    ws3.append(["Total Duration (s)", s["total_duration_s"]])
    autosize(ws3)

    ws4 = wb.create_sheet("Defect Summary")
    header_row(ws4, ["#", "Defect / Test ID", "Module", "Severity"])
    i = 0
    for r in results:
        if r["status"] != "FAILED":
            continue
        i += 1
        severity = MODULE_SEVERITY.get(r["module_name"], "MEDIUM")
        ws4.append([i, short_test_id(r["nodeid"]), r["module_name"], severity])
    autosize(ws4)

    out_path = os.path.join(config.REPORTS_DIR, "Automation_Test_Report.xlsx")
    wb.save(out_path)
    print(f"Wrote {out_path}")


def write_execution_report_html(payload: dict):
    s = payload["summary"]
    results = payload["results"]
    generated_at = payload["generated_at"]

    rows_html = []
    for r in results:
        color = STATUS_COLORS.get(r["status"], "#333")
        rows_html.append(
            f"<tr><td>{r['nodeid']}</td>"
            f"<td style='color:{color};font-weight:600'>{r['status']}</td>"
            f"<td>{r['duration_s']}s</td></tr>"
        )

    html = f"""<!doctype html>
<html><head><meta charset="utf-8">
<title>FitFuel Mobile - Appium Execution Report</title>
<style>
body {{ font-family: -apple-system, Segoe UI, Roboto, sans-serif; margin: 2rem; color: #1a1a1a; }}
h1 {{ margin-bottom: 0.25rem; }}
.meta {{ color: #666; margin-bottom: 1.5rem; }}
.cards {{ display: flex; gap: 1rem; margin-bottom: 2rem; flex-wrap: wrap; }}
.card {{ padding: 1rem 1.5rem; border-radius: 10px; background: #f4f4f4; min-width: 120px; }}
.card b {{ display:block; font-size: 1.6rem; }}
table {{ border-collapse: collapse; width: 100%; }}
th, td {{ text-align: left; padding: 0.5rem 0.75rem; border-bottom: 1px solid #eee; font-size: 0.9rem; }}
th {{ background: #fafafa; position: sticky; top: 0; }}
</style></head>
<body>
<h1>FitFuel Mobile - Appium Execution Report</h1>
<p class="meta">App: {payload['app_under_test']['package']} &middot; Backend: {payload['app_under_test']['backend_base_url']} &middot; Generated: {generated_at}</p>
<div class="cards">
  <div class="card">Total<b>{s['total']}</b></div>
  <div class="card" style="background:#eaf7ee">Passed<b style="color:#1e8e4a">{s['passed']}</b></div>
  <div class="card" style="background:#fdeceb">Failed<b style="color:#d64545">{s['failed']}</b></div>
  <div class="card" style="background:#fdf3d9">Skipped<b style="color:#b58900">{s['skipped']}</b></div>
  <div class="card">Pass rate<b>{s['pass_rate']}%</b></div>
  <div class="card">Duration<b>{s['total_duration_s']}s</b></div>
</div>
<table>
<thead><tr><th>Test</th><th>Status</th><th>Duration</th></tr></thead>
<tbody>
{''.join(rows_html)}
</tbody>
</table>
</body></html>
"""
    out_path = os.path.join(config.REPORTS_DIR, "execution-report.html")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"Wrote {out_path}")


def write_dashboard_html(payload: dict):
    s = payload["summary"]
    generated_at = payload["generated_at"]
    gate_class = "gate-pass" if s["gate_passed"] else "gate-fail"

    bars = []
    for mod, counts in sorted(s["by_module"].items()):
        total_mod = counts.get("total", 0)
        if total_mod == 0:
            continue
        passed_pct = counts.get("passed", 0) / total_mod * 100
        failed_pct = counts.get("failed", 0) / total_mod * 100
        skipped_pct = counts.get("skipped", 0) / total_mod * 100
        bars.append(
            f"""<div style="margin-bottom:0.75rem">
  <div style="font-size:0.85rem;margin-bottom:0.2rem">{mod} ({total_mod})</div>
  <div style="display:flex;height:14px;border-radius:7px;overflow:hidden;background:#eee">
    <div style="width:{passed_pct}%;background:#1e8e4a"></div>
    <div style="width:{failed_pct}%;background:#d64545"></div>
    <div style="width:{skipped_pct}%;background:#b58900"></div>
  </div>
</div>"""
        )

    html = f"""<!doctype html>
<html><head><meta charset="utf-8"><title>FitFuel Mobile - Test Dashboard</title>
<style>
body {{ font-family: -apple-system, Segoe UI, Roboto, sans-serif; margin: 2rem; }}
h1 {{ margin-bottom: 0.25rem; }}
.big {{ font-size: 3rem; font-weight: 700; }}
.gate-pass {{ color: #1e8e4a; }}
.gate-fail {{ color: #d64545; }}
</style></head>
<body>
<h1>FitFuel Mobile - Test Dashboard</h1>
<p class="big {gate_class}">
  {s['pass_rate']}%
</p>
<p>Gate: {s['gate_threshold_pct']}% required to pass CI &middot;
   {s['total']} tests total &middot;
   {s['passed']} passed / {s['failed']} failed / {s['skipped']} skipped &middot;
   {s['total_duration_s']}s total &middot; generated {generated_at}</p>
<h2>By module</h2>
{''.join(bars)}
</body></html>
"""
    out_path = os.path.join(config.REPORTS_DIR, "dashboard.html")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"Wrote {out_path}")


def write_summary_md(payload: dict):
    s = payload["summary"]
    lines = [
        "# FitFuel Mobile - Appium Test Summary",
        "",
        f"- **App package:** {payload['app_under_test']['package']}",
        f"- **Total tests:** {s['total']}",
        f"- **Passed:** {s['passed']}",
        f"- **Failed:** {s['failed']}",
        f"- **Skipped:** {s['skipped']}",
        f"- **Pass rate:** {s['pass_rate']}% (threshold: {s['gate_threshold_pct']}%)",
        f"- **Gate:** {'PASSED' if s['gate_passed'] else 'FAILED'}",
        f"- **Total duration:** {s['total_duration_s']}s",
        "",
        "## By module",
        "",
        "| Module | Passed | Failed | Skipped | Total |",
        "|---|---|---|---|---|",
    ]
    for mod, counts in sorted(s["by_module"].items()):
        lines.append(
            f"| `{mod}` | {counts.get('passed', 0)} | {counts.get('failed', 0)} | "
            f"{counts.get('skipped', 0)} | {counts.get('total', 0)} |"
        )

    out_path = os.path.join(config.REPORTS_DIR, "summary.md")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    print(f"Wrote {out_path}")


def main():
    os.makedirs(config.REPORTS_DIR, exist_ok=True)
    payload = load_results()

    write_excel_report(payload)
    write_execution_report_html(payload)
    write_dashboard_html(payload)
    write_summary_md(payload)

    s = payload["summary"]
    print(
        f"\n=== SUMMARY: total={s['total']} (expected {EXPECTED_TEST_COUNT}) "
        f"passed={s['passed']} failed={s['failed']} skipped={s['skipped']} "
        f"pass_rate={s['pass_rate']}% "
        f"gate={'PASS' if s['gate_passed'] else 'FAIL'} (threshold {s['gate_threshold_pct']}%) ==="
    )

    if not s["gate_passed"]:
        print("Pass-rate gate NOT met.", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()