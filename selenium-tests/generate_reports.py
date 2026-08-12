"""
generate_reports.py

Runs as a SEPARATE step AFTER `pytest` finishes (see the CI workflow). It
merges every reports/results/result_*.json file (one per xdist worker, or a
single result_master.json for a non-distributed run) into:

  reports/execution-results.json      - merged raw + summary, machine-readable
  reports/Automation_Test_Report.xlsx - Executed Tests / Passed / Failed /
                                          Skipped / Execution Metrics / Defect
                                          Summary sheets
  reports/execution-report.html       - full per-test table, one row per attempt
  reports/dashboard.html              - pass-rate gate + per-module bar chart
  reports/summary.md                  - short markdown summary for a PR comment

No process ever writes more than one result_*.json (see conftest.py), so
merging here is a plain concatenation - no de-duplication logic needed.
"""

from __future__ import annotations

import glob
import json
import os
import sys

sys.path.insert(0, os.path.dirname(__file__))

from config import PASS_RATE_GATE, REPORTS_DIR, RESULTS_DIR  # noqa: E402

MODULE_SEVERITY = {
    "Authentication": "HIGH",
    "Authorization": "HIGH",
    "CRUD Operations": "MEDIUM",
    "Forms": "MEDIUM",
    "Input Validation": "MEDIUM",
    "Error Handling": "MEDIUM",
    "Session Management": "MEDIUM",
    "Navigation": "LOW",
    "UI Validation": "LOW",
    "Downloads & Export": "LOW",
    "Accessibility": "LOW",
    "Responsive": "LOW",
}

STATUS_COLORS = {
    "passed": "#1e8e4a",
    "failed": "#d64545",
    "skipped": "#b58900",
    "rerun": "#333333",
}


def load_all_results() -> tuple[list[dict], str]:
    files = sorted(glob.glob(os.path.join(RESULTS_DIR, "result_*.json")))
    if not files:
        print(f"No result files found under {RESULTS_DIR}/ - did pytest run first?", file=sys.stderr)
        sys.exit(1)

    combined: list[dict] = []
    base_url = ""
    generated_ats = []
    for path in files:
        with open(path, encoding="utf-8") as f:
            payload = json.load(f)
        combined.extend(payload.get("results", []))
        base_url = payload.get("base_url", base_url)
        generated_ats.append(payload.get("generated_at", ""))

    generated_at = max(generated_ats) if generated_ats else ""
    print(f"Loaded {len(combined)} result entries from {len(files)} file(s): {[os.path.basename(f) for f in files]}")
    return combined, base_url


def build_summary(results: list[dict], base_url: str) -> dict:
    passed = sum(1 for r in results if r["status"] == "passed")
    failed = sum(1 for r in results if r["status"] == "failed")
    skipped = sum(1 for r in results if r["status"] == "skipped")
    reran = sum(1 for r in results if r["status"] == "rerun")
    total = len(results)
    total_duration = round(sum(r.get("duration_s", 0.0) for r in results), 3)

    denominator = passed + failed
    pass_rate = round((passed / denominator) * 100, 2) if denominator else 0.0

    by_module: dict[str, dict] = {}
    for r in results:
        mod = r["module"]
        entry = by_module.setdefault(mod, {})
        entry[r["status"]] = entry.get(r["status"], 0) + 1

    return {
        "base_url": base_url,
        "summary": {
            "total": total,
            "passed": passed,
            "failed": failed,
            "skipped": skipped,
            "rerun": reran,
            "pass_rate": pass_rate,
            "total_duration_s": total_duration,
            "gate_threshold_pct": PASS_RATE_GATE,
            "gate_passed": pass_rate >= PASS_RATE_GATE,
            "by_module": by_module,
        },
        "results": results,
    }


def write_execution_results_json(summary: dict, generated_at: str):
    summary["generated_at"] = generated_at
    out_path = os.path.join(REPORTS_DIR, "execution-results.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2)
    print(f"Wrote {out_path}")


def short_test_id(nodeid: str) -> str:
    """'tests/test_x.py::TestClass::test_name[param]' -> 'test_name[param]'"""
    return nodeid.split("::")[-1]


def write_excel_report(summary: dict, generated_at: str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    results = summary["results"]
    s = summary["summary"]

    wb = openpyxl.Workbook()

    header_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    header_font = Font(bold=True)
    status_fill = {
        "passed": PatternFill(start_color="EAF7EE", end_color="EAF7EE", fill_type="solid"),
        "failed": PatternFill(start_color="FDECEB", end_color="FDECEB", fill_type="solid"),
        "skipped": PatternFill(start_color="FDF3D9", end_color="FDF3D9", fill_type="solid"),
        "rerun": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
    }

    def autosize(ws):
        for col_cells in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, 10), 90)

    def header_row(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

    # --- Executed Tests (every attempt, including reruns) ---
    ws = wb.active
    ws.title = "Executed Tests"
    header_row(ws, ["#", "Test ID", "Module", "Markers", "Status", "Duration (s)"])
    for i, r in enumerate(results, start=1):
        ws.append([i, short_test_id(r["nodeid"]), r["module_name"], r.get("markers", ""), r["status"].upper(), r["duration_s"]])
        ws.cell(row=i + 1, column=5).fill = status_fill.get(r["status"], PatternFill())
    autosize(ws)

    # --- Passed / Failed / Skipped ---
    for status_key, sheet_name in (("passed", "Passed"), ("failed", "Failed"), ("skipped", "Skipped")):
        ws2 = wb.create_sheet(sheet_name)
        header_row(ws2, ["#", "Test ID", "Module", "Duration (s)"])
        rows = [r for r in results if r["status"] == status_key]
        for i, r in enumerate(rows, start=1):
            ws2.append([i, short_test_id(r["nodeid"]), r["module_name"], r["duration_s"]])
        autosize(ws2)

    # --- Execution Metrics ---
    ws3 = wb.create_sheet("Execution Metrics")
    header_row(ws3, ["Metric", "Value"])
    ws3.append(["Run At", generated_at])
    ws3.append(["Base URL", summary["base_url"]])
    ws3.append(["Total Tests", s["total"]])
    ws3.append(["Passed", s["passed"]])
    ws3.append(["Failed", s["failed"]])
    ws3.append(["Skipped", s["skipped"]])
    ws3.append(["Reruns", s["rerun"]])
    ws3.append(["Pass Rate (%)", s["pass_rate"]])
    ws3.append(["Gate Threshold (%)", s["gate_threshold_pct"]])
    ws3.append(["Gate Passed", "YES" if s["gate_passed"] else "NO"])
    ws3.append(["Total Duration (s)", s["total_duration_s"]])
    autosize(ws3)

    # --- Defect Summary (failed tests, deduplicated by test id) ---
    ws4 = wb.create_sheet("Defect Summary")
    header_row(ws4, ["#", "Defect / Test ID", "Module", "Severity"])
    seen = set()
    i = 0
    for r in results:
        if r["status"] != "failed":
            continue
        key = r["nodeid"]
        if key in seen:
            continue
        seen.add(key)
        i += 1
        severity = MODULE_SEVERITY.get(r["module_name"], "MEDIUM")
        ws4.append([i, short_test_id(r["nodeid"]), r["module_name"], severity])
    autosize(ws4)

    out_path = os.path.join(REPORTS_DIR, "Automation_Test_Report.xlsx")
    wb.save(out_path)
    print(f"Wrote {out_path}")


def write_execution_report_html(summary: dict, generated_at: str):
    s = summary["summary"]
    results = summary["results"]

    rows_html = []
    for r in results:
        color = STATUS_COLORS.get(r["status"], "#333")
        rows_html.append(
            f"<tr><td>{r['nodeid']}</td>"
            f"<td style='color:{color};font-weight:600'>{r['status'].upper()}</td>"
            f"<td>{r['duration_s']}s</td></tr>"
        )

    html = f"""<!doctype html>
<html><head><meta charset="utf-8">
<title>FitFuel - Selenium Execution Report</title>
<style>
body {{ font-family: -apple-system, Segoe UI, Roboto, sans-serif; margin: 2rem; color: #1a1a1a; }}
h1 {{ margin-bottom: 0.25rem; }}
.meta {{ color: #666; margin-bottom: 1.5rem; }}
.cards {{ display: flex; gap: 1rem; margin-bottom: 2rem; flex-wrap: wrap; }}
.card {{ padding: 1rem 1.5rem; border-radius: 10px; background: #f4f4f4; min-width: 120px; }}
.card b {{ display:block; font-size: 1.6rem; }}
table {{ border-collapse: collapse; width: 100%; }}
th, td {{ text-align: left; padding: 0.5rem 0.75rem; border-bottom: 1px solid #eee; font-size: 0.9rem; }}
th {{ background: #fafafa; position: sticky; top: 0; }}
</style></head>
<body>
<h1>FitFuel - Selenium Execution Report</h1>
<p class="meta">Target: {summary['base_url']} &middot; Generated: {generated_at}</p>
<div class="cards">
  <div class="card">Total<b>{s['total']}</b></div>
  <div class="card" style="background:#eaf7ee">Passed<b style="color:#1e8e4a">{s['passed']}</b></div>
  <div class="card" style="background:#fdeceb">Failed<b style="color:#d64545">{s['failed']}</b></div>
  <div class="card" style="background:#fdf3d9">Skipped<b style="color:#b58900">{s['skipped']}</b></div>
  <div class="card">Reruns<b>{s['rerun']}</b></div>
  <div class="card">Pass rate<b>{s['pass_rate']}%</b></div>
  <div class="card">Duration<b>{s['total_duration_s']}s</b></div>
</div>
<table>
<thead><tr><th>Test</th><th>Status</th><th>Duration</th></tr></thead>
<tbody>
{''.join(rows_html)}
</tbody>
</table>
</body></html>
"""
    out_path = os.path.join(REPORTS_DIR, "execution-report.html")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"Wrote {out_path}")


def write_dashboard_html(summary: dict, generated_at: str):
    s = summary["summary"]
    gate_class = "gate-pass" if s["gate_passed"] else "gate-fail"

    bars = []
    for mod, counts in sorted(s["by_module"].items()):
        total_mod = sum(counts.values())
        if total_mod == 0:
            continue
        passed_pct = counts.get("passed", 0) / total_mod * 100
        failed_pct = counts.get("failed", 0) / total_mod * 100
        skipped_pct = counts.get("skipped", 0) / total_mod * 100
        mod_label = os.path.basename(mod).replace(".py", "")
        bars.append(
            f"""<div style="margin-bottom:0.75rem">
  <div style="font-size:0.85rem;margin-bottom:0.2rem">{mod_label} ({total_mod})</div>
  <div style="display:flex;height:14px;border-radius:7px;overflow:hidden;background:#eee">
    <div style="width:{passed_pct}%;background:#1e8e4a"></div>
    <div style="width:{failed_pct}%;background:#d64545"></div>
    <div style="width:{skipped_pct}%;background:#b58900"></div>
  </div>
</div>"""
        )

    html = f"""<!doctype html>
<html><head><meta charset="utf-8"><title>FitFuel - Test Dashboard</title>
<style>
body {{ font-family: -apple-system, Segoe UI, Roboto, sans-serif; margin: 2rem; }}
h1 {{ margin-bottom: 0.25rem; }}
.big {{ font-size: 3rem; font-weight: 700; }}
.gate-pass {{ color: #1e8e4a; }}
.gate-fail {{ color: #d64545; }}
</style></head>
<body>
<h1>FitFuel - Test Dashboard</h1>
<p class="big {gate_class}">
  {s['pass_rate']}%
</p>
<p>Gate: {s['gate_threshold_pct']}% required to pass CI &middot;
   {s['passed']} passed / {s['failed']} failed / {s['skipped']} skipped &middot;
   {s['total_duration_s']}s total</p>
<h2>By module</h2>
{''.join(bars)}
</body></html>
"""
    out_path = os.path.join(REPORTS_DIR, "dashboard.html")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"Wrote {out_path}")


def write_summary_md(summary: dict):
    s = summary["summary"]
    lines = [
        "# FitFuel - Selenium Web Test Summary",
        "",
        f"- **Total executed:** {s['total']}",
        f"- **Passed:** {s['passed']}",
        f"- **Failed:** {s['failed']}",
        f"- **Skipped:** {s['skipped']}",
        f"- **Reruns:** {s['rerun']}",
        f"- **Pass rate:** {s['pass_rate']}% (threshold: {s['gate_threshold_pct']}%)",
        f"- **Gate:** {'PASSED' if s['gate_passed'] else 'FAILED'}",
        f"- **Total duration:** {s['total_duration_s']}s",
        "",
        "## By module",
        "",
        "| Module | Passed | Failed | Skipped |",
        "|---|---|---|---|",
    ]
    for mod, counts in sorted(s["by_module"].items()):
        lines.append(f"| `{mod}` | {counts.get('passed', 0)} | {counts.get('failed', 0)} | {counts.get('skipped', 0)} |")

    out_path = os.path.join(REPORTS_DIR, "summary.md")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    print(f"Wrote {out_path}")


def main():
    os.makedirs(REPORTS_DIR, exist_ok=True)
    results, base_url = load_all_results()
    summary = build_summary(results, base_url)

    import time as _time

    generated_at = _time.strftime("%Y-%m-%dT%H:%M:%S", _time.gmtime()) + "+00:00"

    write_execution_results_json(summary, generated_at)
    write_excel_report(summary, generated_at)
    write_execution_report_html(summary, generated_at)
    write_dashboard_html(summary, generated_at)
    write_summary_md(summary)

    s = summary["summary"]
    print(
        f"\n=== SUMMARY: total={s['total']} passed={s['passed']} failed={s['failed']} "
        f"skipped={s['skipped']} rerun={s['rerun']} pass_rate={s['pass_rate']}% "
        f"gate={'PASS' if s['gate_passed'] else 'FAIL'} (threshold {s['gate_threshold_pct']}%) ==="
    )

    if not s["gate_passed"]:
        print("Pass-rate gate NOT met.", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
